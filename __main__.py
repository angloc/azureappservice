from azure.identity import DefaultAzureCredential
from azure.graphrbac import GraphRbacManagementClient
from azure.mgmt.resource import SubscriptionClient
#from azure.mgmt.keyvault import KeyVaultManagementClient
#from azure.mgmt.keyvault.models import AccessPolicyEntry, Permissions
#from azure.mgmt.storage import StorageManagementClient
#from azure.mgmt.storage.models import StorageAccountCreateParameters, StorageAccountUpdateParameters
from azure.core.exceptions import ClientAuthenticationError

import pulumi
import pulumi_azure as azure

import openpyxl

# Function to convert rows to dictionaries based on the header row
def rows_to_dicts(sheet):
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    return [
        {headers[i]: cell.value for i, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

def worksheet_to_dict(worksheet):
    data_dict = {}
    for row in worksheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key, value = (c.strip() if c else None for c in row)
        if key and value:  # Ensuring the key is not None
            data_dict[key] = value
    return data_dict

def validate_azure (tenant_name, subscription_name):
    # Authenticate with Azure

    credential = DefaultAzureCredential()
    subscription_client = SubscriptionClient(credential)

    try:
        # Get subscription details
        subscription_details = subscription_client.subscriptions.list()
        known_subscriptions = {}
        for subscription in subscription_details:
            name = subscription.display_name
            tenant_id = subscription.tenant_id
            print(f"Subscription {name} belongs to {known_tenants_by_id [tenant.id].display_name}")
            known_subscriptions [name] = subscription
    except ClientAuthenticationError:
        raise ValueError("Authentication failed - ensure you are logged in to Azure CLI")

    if subscription_name not in known_subscriptions:
        raise ValueError(f"Subscription {subscription_name} not found in your credentials")

    # Initialize the GraphRbacManagementClient
    graph_client = GraphRbacManagementClient(credential, tenant_id)

    # Get the tenant details
    known_tenants_by_name = {}
    known_tenants_by_id = {}

    tenant_details = graph_client.tenants.list()
    for tenant in tenant_details:
        print(f"Tenant ID: {tenant.tenant_id}")
        print(f"Tenant Name: {tenant.display_name}")
        known_tenants_by_name [tenant.display_name] = tenant
        known_tenants_by_id [tenant.id] = tenant

    # Validate tenant
    if tenant_name not in known_tenants_by_name:
        raise ValueError(f"Tenant {tenant} not found")
    else:
        tenant_id = known_tenants_by_name[tenant].tenant_id

    # Validate subscription
    if subscription not in known_subscriptions:
        raise ValueError(f"Subscription {subscription} not found")
    else:
        subscription_id = known_subscriptions[subscription].subscription_id

    return tenant, subscription

def validate_resources(config_file):
    # Load the workbook
    workbook = openpyxl.load_workbook(config_file)
    # Read Pulumi worksheet
    pulumi_sheet = workbook['Configuration']
    config = worksheet_to_dict(pulumi_sheet)
    if 'Tenant' not in config:
        raise ValueError("Tenant not found in Config worksheet")
    tenant_name = config['Tenant']
    if 'Subscription' not in config:
        raise ValueError("Subscription not found in Config worksheet")
    subscription_name = config['Subscription']
    if 'Pulumi Resource Group' not in config:
        raise ValueError("'Pulumi Resource Group' not found in Config worksheet")
    pulumi_resource_group = config['Pulumi Resource Group']
    pulumi_storage_account = config.get("Pulumi Storage Account", "pulumi")
    pulumi_container = config.get("Pulumi Container", "pulumi")
    tenant, subscription = validate_azure(tenant_name, subscription_name)
    return tenant, subscription, pulumi_resource_group, pulumi_storage_account, pulumi_container

def deploy_resources():

    config = pulumi.Config()
    config_file = config.require('config-file', )

    # Authenticate with Azure

    credential = DefaultAzureCredential()
    subscription_client = SubscriptionClient(credential)


    # Initialize the GraphRbacManagementClient
    graph_client = GraphRbacManagementClient(credential, tenant_id)

    # Get the tenant details
    known_tenants_by_name = {}
    known_tenants_by_id = {}
    tenant_details = graph_client.tenants.list()
    for tenant in tenant_details:
        print(f"Tenant ID: {tenant.tenant_id}")
        print(f"Tenant Name: {tenant.display_name}")
        known_tenants_by_name [tenant.display_name] = tenant
        known_tenants_by_id [tenant.id] = tenant

    # Get subscription details
    subscription_details = subscription_client.subscriptions.list()
    known_subscriptions = {}
    for subscription in subscription_details:
        subscription_name = subscription.display_name
        tenant_id = subscription.tenant_id
        print(f"Subscription {subscription_name} belongs to {known_tenants_by_id [tenant.id].display_name}")
        known_subscriptions [subscription_name] = subscription

    # Load the workbook
    workbook = openpyxl.load_workbook(config_file)


    return

    # Read deployments worksheet

    deployments_sheet = workbook['Deployments']
    deployments = rows_to_dicts(deployments_sheet)

    for deployment in deployments:
        resource_group_name = deployment['Resource Group']
        required_tenant_name = deployment['Tenant']
        requrired_subscription_name = deployment['Subscription']
        service_name = deployment['Service']
        app_name = deployment['App']
        deployment_name = f"service-name-"

        # Create an App Service Plan
        app_service_plan = azure.appservice.Plan(
            service_name,
            resource_group_name=resource_group.name,
            location=deployment['Region'],
            sku={'tier': deployment['Tier'], 'size': deployment['Size']}
        )

        # Create a Storage Account
        storage_account = azure.storage.Account(
            f"{service_name}storage",
            resource_group_name=resource_group.name,
            account_replication_type="LRS",
            account_tier="Standard"
        )

        # Create an Azure File Share
        file_share = azure.storage.Share(
            f"{service_name}fileshare",
            storage_account_name=storage_account.name,
            resource_group_name=resource_group.name,
            quota=50
        )

        # Create an Application Insights instance
        app_insights = azure.appinsights.Insights(
            "myAppInsights",
            resource_group_name=resource_group.name,
            application_type="web"
        )

        # Create a Key Vault
        key_vault = azure.keyvault.KeyVault(
            f"{service_name}keyvault",
            resource_group_name=resource_group.name,
            sku_name="standard",
            tenant_id=deployment['TenantId']
        )

        # Create an App Service with a system-assigned managed identity
        app_service = azure.appservice.AppService(
            app_name,
            resource_group_name=resource_group.name,
            app_service_plan_id=app_service_plan.id,
            app_settings={
                "WEBSITE_STOPPED": "1" if deployment['Status'] == 'stopped' else "0",
                "APPINSIGHTS_INSTRUMENTATIONKEY": app_insights.instrumentation_key
            },
            identity=azure.appservice.AppServiceIdentityArgs(type="SystemAssigned")
        )

        # Assign access policy to the Key Vault for the managed identity
        keyvault.AccessPolicy(
            f"{app_name}-access",
            key_vault_id=key_vault.id,
            tenant_id=deployment['TenantId'],
            object_id=app_service.identity.apply(lambda id: id.principal_id),
            key_permissions=["get"],
            secret_permissions=["get"]
        )

if __name__ == "__main__":
    import argparse
    # Set up argument parsing
    parser = argparse.ArgumentParser(description="Deploy Azure resources based on an Excel configuration")
    parser.add_argument("configFile", help="Path to the Excel configuration file")
    args = parser.parse_args()

    # When called outside Pulumi
    validate_resources(args.configFile)
else:
    deploy_resources()
