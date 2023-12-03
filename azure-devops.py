import enum
import os
import subprocess
import tempfile
import sys

from azure.identity import DefaultAzureCredential
from azure.graphrbac import GraphRbacManagementClient
from azure.mgmt.resource import SubscriptionClient
from azure.mgmt.keyvault import KeyVaultManagementClient
from azure.mgmt.keyvault.models import AccessPolicyEntry, Permissions
from azure.mgmt.resource import ResourceManagementClient
from azure.mgmt.storage import StorageManagementClient
from azure.mgmt.storage.models import StorageAccountCreateParameters, StorageAccountUpdateParameters
from azure.storage.blob import BlobServiceClient
from azure.core.exceptions import ClientAuthenticationError, ResourceExistsError

import pulumi
import pulumi_azure
from pulumi.automation import LocalWorkspace, LocalWorkspaceOptions, Stack, ProjectSettings, select_stack

import openpyxl

class ResourceTypes(enum.Enum):
    APP_SERVICE_PLAN = "app_service_plan"
    STORAGE_ACCOUNT = "storage_account"
    FILE_SHARE = "file_share"
    APP_INSIGHTS = "app_insights"
    KEY_VAULT = "key_vault"
    APP_SERVICE = "app_service"
    RESOURCE_GROUP = "resource_group"
    DATABASE_SERVER = "database_server"
    MANAGED_IDENTITY = "managed_identity"
    DATABASE = "database"

defaultTemplates = {
    ResourceTypes.APP_SERVICE_PLAN: "{Service}",
    ResourceTypes.APP_SERVICE: "{Subscription}-{Service}-{App}",
    ResourceTypes.APP_INSIGHTS: "{Service}",
    ResourceTypes.FILE_SHARE: "files",
    ResourceTypes.KEY_VAULT: "{Subscription}-{Service}",
    ResourceTypes.STORAGE_ACCOUNT: "{Subscription}{Service}",
    ResourceTypes.DATABASE_SERVER: "{Subscription}-{Service}",
    ResourceTypes.DATABASE: "{App}",
    ResourceTypes.MANAGED_IDENTITY: "{Subscription}{Service}",
}

# Function to convert rows to dictionaries based on the header row
def rows_to_dicts(sheet):
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    return [
        {headers[i]: cell for i, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

def worksheet_to_dict(worksheet):
    data_dict = {}
    for row in worksheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key, value = (c.strip() if c else None for c in row)
        if key and value:  # Ensuring the key is not None
            data_dict[key] = value
    return data_dict

def validate_azure (subscription_name):
    # Authenticate with Azure

    credential = DefaultAzureCredential()
    subscription_client = SubscriptionClient(credential)

    try:
        # Get subscription details
        subscription_details = subscription_client.subscriptions.list()
        known_subscriptions = {}
        for subscription in subscription_details:
            name = subscription.display_name
            tenant_id = subscription.tenant_id
            known_subscriptions [name] = subscription
    except ClientAuthenticationError:
        raise ValueError("Authentication failed - ensure you are logged in to Azure CLI")

    if subscription_name not in known_subscriptions:
        raise ValueError(f"Subscription {subscription_name} not found in your credentials")
    subscription = known_subscriptions [subscription_name]
    print (f"Using Azure subscription '{subscription_name}' ({subscription.subscription_id})")

    return subscription

def validate_resources(config_file):
    # Load the workbook
    workbook = openpyxl.load_workbook(config_file)
    # Read Pulumi worksheet
    pulumi_sheet = workbook['Configuration']
    config = worksheet_to_dict(pulumi_sheet)
    templates = {}
    templates.update(defaultTemplates)
    if 'Templates' in workbook:
        templates_sheet = workbook['Templates']
        templates.update (rows_to_dicts(templates_sheet))
    workbook.close()
    for name, value in templates.items():
        print (f"{name}: {value}")
    if 'Subscription' not in config:
        raise ValueError("Subscription not found in Configuration worksheet")
    subscription_name = config['Subscription']
    if 'Subscription slug' not in config:
        raise ValueError("Subscription slug not found in Configuration worksheet")
    subscription_slug = config['Subscription slug'].lower()
    if 'Pulumi Resource Group' not in config:
        raise ValueError("'Pulumi Resource Group' not found in Configuration worksheet")
    pulumi_resource_group = config['Pulumi Resource Group'].lower()
    if 'Pulumi Storage Account' not in config:
        raise ValueError("'Pulumi Storage Account' not found in Configuration worksheet")
    pulumi_storage_account = config.get("Pulumi Storage Account", "pulumi").lower()
    pulumi_container = config.get("Pulumi Container", "pulumi").lower()
    pulumi_location = config.get("Pulumi Location", "uksouth").lower()
    subscription = validate_azure(subscription_name)

    return templates, subscription, subscription_slug, pulumi_resource_group, pulumi_storage_account, pulumi_location, pulumi_container

def send_script_to_pulumi(script, initialiser=""):
    if initialiser: script = initialiser + "\n" + script
    #print ("Sending to Pulumi", script)
    with tempfile.NamedTemporaryFile(mode="w", delete=False) as f:
        f.write(script)
        f.flush()
        subprocess.run(["bash", f.name], check=True)
        os.remove (f.name)

def ensure_pulumi_resources(subscription_id, resource_group_name, storage_account_name, location, container_name, stack_name):

    # Create a credential object using DefaultAzureCredential
    credential = DefaultAzureCredential()

    # Create a client object for Resource Management
    resource_client = ResourceManagementClient(credential, subscription_id)

    # Create a client object for Storage Management
    storage_client = StorageManagementClient(credential, subscription_id)

    # Check if the resource group exists, and create if not
    resource_group = resource_client.resource_groups.check_existence(resource_group_name)
    if not resource_group:
        print(f"Creating resource group '{resource_group_name}'at location '{location}'")
        resource_client.resource_groups.create_or_update(resource_group_name, {'location': location})

    # Check if the storage account exists, and create if not
    storage_accounts = storage_client.storage_accounts.list_by_resource_group(resource_group_name)
    storage_account_exists = any(account.name == storage_account_name for account in storage_accounts)

    if not storage_account_exists:
        print(f"Creating storage account '{storage_account_name}'...")
        try:
            storage_client.storage_accounts.begin_create(resource_group_name, storage_account_name, {
                'location': location,
                'sku': {'name': 'Standard_LRS'},
                'kind': 'StorageV2'
            }).result()
        except ResourceExistsError as e:
            print(f"Storage account '{storage_account_name}' already exists elsewhere")

    # Get the storage account key
    storage_keys = storage_client.storage_accounts.list_keys(resource_group_name, storage_account_name)
    storage_key = storage_keys.keys[0].value
    storage_url = f"https://{storage_account_name}.blob.core.windows.net"

    # Create a BlobServiceClient to manage containers
    blob_service_client = BlobServiceClient(account_url=storage_url, credential=storage_key)

    # Check if the container exists, and create if not
    container_client = blob_service_client.get_container_client(container_name)
    if not container_client.exists():
        print(f"Creating Pulumi container '{container_name}'...")
        blob_service_client.create_container(container_name)

    # Pulumi stack file name
    stack_file_name = f".pulumi/stacks/devops/{stack_name}.json"

    PULUMI_INITIALISATION_SCRIPT = f"""
        #!/bin/bash
        set -e
        export AZURE_STORAGE_ACCOUNT={storage_account_name}
        export AZURE_STORAGE_KEY={storage_key}
        export CONFIG_FILE=f"{stack_name}.xlsx"
        export PULUMI_CONFIG_PASSPHRASE="No secrets here"
        pulumi login --cloud-url azblob://{container_name}
    """

    # Check if Pulumi project has been set up
    blob_list = list(container_client.list_blobs())
    project_exists = any(blob.name.startswith('.pulumi') for blob in blob_list)
    if not project_exists:
        print(f"Creating Pulumi project 'devops'")
        send_script_to_pulumi(
            "pulumi new azure-python --name devops --yes",
            PULUMI_INITIALISATION_SCRIPT
        )

    # Check if Pulumi stack for this spreadsheet has been set up
    stack_exists = any(blob.name == stack_file_name for blob in blob_list)
    if not stack_exists:       
        # Create a new stack
        print(f"Creating Pulumi stack {stack_name}")
        send_script_to_pulumi(f"pulumi stack init {stack_name}", PULUMI_INITIALISATION_SCRIPT)
    print(f"All Azure resources for Pulumi stack '{stack_name}' are set up in storage account '{storage_account_name}'")
    return storage_url, storage_key, PULUMI_INITIALISATION_SCRIPT

def upload_file_to_blob(account_url, account_key, container_name, file_path):
    try:
        # Create a BlobServiceClient using the storage account name and account key
        blob_service_client = BlobServiceClient(account_url, credential=account_key)
        
        # Get a reference to the container
        container_client = blob_service_client.get_container_client(container_name)

        # Get the file name from the file path
        local_file_name = os.path.basename(file_path)

        # Create a blob client using the local file name as the name for the blob
        blob_client = container_client.get_blob_client(blob=local_file_name)

        # Upload the local file to the blob
        with open(file_path, "rb") as data:
            blob_client.upload_blob(data, overwrite=True)

        print(f"..File {local_file_name} uploaded to blob container {container_name}")

    except Exception as e:
        print(f'--Failed to upload {file_path} to Pulumi container: {e}')

def deploy_resources(config_file):

    stack_name = os.path.splitext(os.path.basename(config_file))[0]

    templates, subscription, subscription_slug, pulumi_resource_group, \
            pulumi_storage_account, pulumi_location, pulumi_container =\
            validate_resources(config_file)
    ensure_pulumi_resources(subscription.subscription_id,
        pulumi_resource_group, pulumi_storage_account, pulumi_location, pulumi_container, stack_name)

    # Read deployments worksheet
    workbook = openpyxl.load_workbook(config_file)

    deployments_sheet = workbook['Deployments']
    deployments = rows_to_dicts(deployments_sheet)
    workbook.close()

    # Advise Pulumi of the current configuration

    def validate_deployments_column(name, index, default=None):
        if not name in deployments [index] or\
            not deployments [index][name] or\
            not deployments [index][name].strip():
            if default:
                deployments [index][name] = default
            else:
                raise ValueError(f"'{name}' is missing from row {index+1} of the deployments worksheet")

    # Create a credential object using DefaultAzureCredential
    credential = DefaultAzureCredential()

    # Create a client object for Resource Management
    resource_client = ResourceManagementClient(credential, subscription.subscription_id)

    resource_groups = {}

    for i, deployment in enumerate (deployments):
        validate_deployments_column('Defines', i)
        validate_deployments_column('Resource Group', i)
        validate_deployments_column('Service', i)
        validate_deployments_column('sku', i, "B1")
        validate_deployments_column('Region', i, "uksouth")
        validate_deployments_column('Files quota', i, "50")
        deployment['Subscription'] = subscription_slug.lower ()

        # Check if the resource group exists
        resource_group_name = deployment['Resource Group']
        resource_group_exists = resource_client.resource_groups.check_existence(resource_group_name)
        if resource_group_exists:
            resource_group = resource_client.resource_groups.get(resource_group_name)
        else:
            raise ValueError(f"Resource group '{resource_group_name}' does not exist in subscription '{subscription.name}'")
        resource_groups [resource_group_name] = resource_group

    for i, deployment in enumerate (deployments):

        resource_group_name = deployment['Resource Group']
        resource_group = resource_groups [resource_group_name]
        service = deployment['Service'] = deployment['Service'].lower()        
        location = deployment['Region'] = deployment['Region'].lower()
        app_name = deployment.get ("App")
        quota = deployment.get ("Files quota")
        sku = deployment['sku']

        defines = deployment['Defines'].lower()
        print (deployment)

        # Create an App Service Plan
        if defines == "service":

            app_service_plan_name = templates [ResourceTypes.APP_SERVICE_PLAN].format (**deployment)
            app_service_plan = pulumi_azure.appservice.ServicePlan(
                app_service_plan_name, name=app_service_plan_name,
                resource_group_name=resource_group_name,
                location=location,
                os_type="Linux",
                sku_name=sku
            )

            # Create a Storage Account
            storage_account_name = templates [ResourceTypes.STORAGE_ACCOUNT].format (**deployment)
            storage_account = pulumi_azure.storage.Account(
                storage_account_name, name=storage_account_name,
                resource_group_name=resource_group_name,
                account_replication_type="LRS",
                account_tier="Standard"
            )

            # Create an Azure File Share
            file_share_name = templates [ResourceTypes.FILE_SHARE].format (**deployment)
            file_share = pulumi_azure.storage.Share(
                file_share_name, name=file_share_name,
                storage_account_name=storage_account.name,
                quota=int(quota)
            )

            # Create an Application Insights instance
            app_insights_name = templates [ResourceTypes.APP_INSIGHTS].format (**deployment)
            app_insights = pulumi_azure.appinsights.Insights(
                app_insights_name, name=app_insights_name,
                resource_group_name=resource_group_name,
                application_type="web"
            )

            # Create a Key Vault
            key_vault_name = templates [ResourceTypes.KEY_VAULT].format (**deployment)
            key_vault = pulumi_azure.keyvault.KeyVault(
                key_vault_name, name=key_vault_name,
                resource_group_name=resource_group_name,
                sku_name="standard",
                tenant_id=subscription.tenant_id,
            )
        
        if defines == "app":

            # Create an App Service with a system-assigned managed identity
            identity=pulumi_azure.appservice.AppServiceIdentityArgs(type="SystemAssigned")
            app_service = pulumi_azure.appservice.AppService(
                app_name,
                resource_group_name=resource_group.name,
                app_service_plan_id=app_service_plan.id,
                app_settings={
                    "WEBSITE_STOPPED": "1" if deployment['Status'] == 'stopped' else "0",
                    "APPINSIGHTS_INSTRUMENTATIONKEY": app_insights.instrumentation_key
                },
                identity=identity
            )

            # Assign access policy to the Key Vault for the managed identity
            key_vault.AccessPolicy(
                f"{app_name}-access",
                key_vault_id=key_vault.id,
                tenant_id=subscription.tenant_id,
                object_id=app_service.identity.apply(lambda id: id.principal_id),
                key_permissions=["get"],
                secret_permissions=["get"]
            )

if __name__ == "__main__":
    import argparse
    import json

    parser = argparse.ArgumentParser(description="Deploy Azure resources based on an Excel configuration")
    parser.add_argument("configFile", help="Path to the Excel configuration file")
    parser.add_argument('--pulumi', type=str, help='Pulumi command to run.')
    args = parser.parse_args()

    templates, subscription, subscription_slug, pulumi_resource_group, pulumi_storage_account, pulumi_location, pulumi_container =\
        validate_resources(args.configFile)
    stack_name = os.path.splitext(os.path.basename(args.configFile))[0]
    storage_url, storage_key, initialiser = ensure_pulumi_resources(subscription.subscription_id,
        pulumi_resource_group, pulumi_storage_account, pulumi_location, pulumi_container, stack_name)

    if args.pulumi:
        send_script_to_pulumi(f"pulumi stack select {stack_name}\npulumi {args.pulumi}", initialiser)
    else:
        os.environ ["AZURE_STORAGE_ACCOUNT"] = pulumi_storage_account
        os.environ ["AZURE_STORAGE_KEY"] = storage_key
        os.environ ["PULUMI_CONFIG_PASSPHRASE"] ="No secrets here"
        project_settings=ProjectSettings(
            name="devops",
            runtime="python",
            backend={"url": f"azblob://{pulumi_container}"}
        )
        workspace = LocalWorkspace(project_settings=project_settings)
        selected_stack = select_stack(
            stack_name=stack_name,
            program=lambda: deploy_resources (args.configFile),
            project_name="devops",
            opts=LocalWorkspaceOptions(project_settings=project_settings)
        )

        up_result = selected_stack.up()
        print(f"update summary: \n{json.dumps(up_result.summary.resource_changes, indent=4)}")

        upload_file_to_blob(storage_url, storage_key, pulumi_container, args.configFile)


