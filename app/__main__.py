import pulumi
import pulumi_azure as azure
import openpyxl
import argparse

config = pulumi.Config()
resource_group_name = config.require('resource_group')
resource_group = azure.core.ResourceGroup(resource_group_name)

def deploy_resources(config_file):
    # Load the workbook
    workbook = openpyxl.load_workbook(config_file)

    # Function to convert rows to dictionaries based on the header row
    def rows_to_dicts(sheet):
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        return [
            {headers[i]: cell.value for i, cell in enumerate(row)}
                for row in sheet.iter_rows(min_row=2, values_only=True)
        ]

    # Read worksheets
    services_sheet = workbook['Services']
    services = rows_to_dicts(services_sheet)

    deployments_sheet = workbook['Deployments']
    deployments = rows_to_dicts(deployments_sheet)

    for deployment in deployments:
        service_name = deployment['Service']
        app_name = deployment['App']

        # Create an App Service Plan
        app_service_plan = azure.appservice.Plan(service_name,
                                                resource_group_name=resource_group.name,
                                                location=deployment['Region'],
                                                sku={'tier': deployment['Tier'], 'size': deployment['Size']})

        # Create a Storage Account
        storage_account = azure.storage.Account(f"{service_name}storage",
                                                resource_group_name=resource_group.name,
                                                account_replication_type="LRS",
                                                account_tier="Standard")

        # Create an Azure File Share
        file_share = azure.storage.Share(f"{service_name}fileshare",
                                        storage_account_name=storage_account.name,
                                        resource_group_name=resource_group.name,
                                        quota=50)

        # Create a Key Vault
        key_vault = azure.keyvault.KeyVault(f"{service_name}keyvault",
                                            resource_group_name=resource_group.name,
                                            sku_name="standard",
                                            tenant_id=deployment['TenantId'])

        # Create an App Service with a system-assigned managed identity
        app_service = azure.appservice.AppService(app_name,
                                                resource_group_name=resource_group.name,
                                                app_service_plan_id=app_service_plan.id,
                                                app_settings={
                                                    "WEBSITE_STOPPED": "1" if deployment['Status'] == 'stopped' else "0"
                                                },
                                                identity=azure.appservice.AppServiceIdentityArgs(type="SystemAssigned"))

        # Assign access policy to the Key Vault for the managed identity
        keyvault.AccessPolicy(f"{app_name}-access",
                            key_vault_id=key_vault.id,
                            tenant_id=deployment['TenantId'],
                            object_id=app_service.identity.apply(lambda id: id.principal_id),
                            key_permissions=["get"],
                            secret_permissions=["get"])

if __name__ == "__main__":
    # Set up argument parsing
    parser = argparse.ArgumentParser(description="Deploy Azure resources based on an Excel configuration")
    parser.add_argument('--config_file', default='config.xlsx', help='Path to the Excel configuration file')
    args = parser.parse_args()

    # Call the deploy function
    deploy_resources(args.config_file)
